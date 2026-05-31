"""Fix unit handling in the PinkSheet loader.

Changes:
  - Oil: $/bbl native, NO conversion (matches load_gas_prices and load_consolidated_prices).
  - Natural Gas: $/MMBtu native, NO conversion.
  - Coal: $/tonne, already correct.
  - Base metals (Cu, Al, Ni, Zn, Pb, Sn): $/tonne, already correct.
  - Precious metals (Gold, Silver, Platinum): $/troy oz × 32,150.7 -> $/tonne.
    Matches consolidated file's $/tonne convention and EI_Prices post-conversion.
  - Iron ore: REMOVED. $/dmtu does not convert cleanly to $/tonne of ore
    (would require iron-content assumption). Also not in Leo's metals list.
"""
import json, sys

NB = '/Users/leoss/Desktop/GitHub/leoss14.github.io/projects/capstone/New code/extension_2024/e0_NR_extraction.ipynb'

NEW_LOADER_SOURCE = '''"""
World Bank Commodity Price Data (the "Pink Sheet").

Annual nominal USD prices, 1960-present, for crude oil, coal, natural gas,
LNG, plus base and precious metals. Single-sheet long-form output matching
the schema produced by the other source loaders.

Download: https://www.worldbank.org/en/research/commodity-markets
File: CMO-Historical-Data-Annual.xlsx (place in rawdata/).

Units returned match the rest of the pipeline:
  Oil          $/bbl       (native; same as load_gas_prices and load_consolidated_prices)
  Natural Gas  $/MMBtu     (native; same as above)
  Coal         $/tonne     (native; same as above)
  Base metals  $/tonne     (native)
  Precious     $/tonne     (converted from $/troy oz x 32,150.7; matches consolidated file)
"""
import openpyxl

PINK_SHEET_PATH = RAW / 'CMO-Historical-Data-Annual.xlsx'

# Map Pink Sheet column header -> (canonical Resource name, native unit).
# Pink Sheet uses US spelling "Aluminum"; canonical is UK "Aluminium".
# Iron ore is skipped: $/dmtu does not convert cleanly to $/tonne of ore
# without an iron-content assumption, and Iron ore is not in the metals list.
PINK_SHEET_MAP = {
    'Crude oil, average':           ('Oil',            '$/bbl'),
    'Coal, Australian':             ('Coal',           '$/mt'),
    'Natural gas, US':              ('Natural Gas',    '$/mmbtu'),
    'Aluminum':                     ('Aluminium',      '$/mt'),
    'Copper':                       ('Copper',         '$/mt'),
    'Lead':                         ('Lead',           '$/mt'),
    'Tin':                          ('Tin',            '$/mt'),
    'Nickel':                       ('Nickel',         '$/mt'),
    'Zinc':                         ('Zinc',           '$/mt'),
    'Gold':                         ('Gold',           '$/troy oz'),
    'Platinum':                     ('Platinum Group', '$/troy oz'),
    'Silver':                       ('Silver',         '$/troy oz'),
}

# Conversion factor to the pipeline's canonical unit for that resource.
# Oil, Natural Gas, Coal: native units ARE canonical -> factor 1.0.
# Base metals: native $/mt IS canonical -> 1.0.
# Precious metals: $/troy oz -> $/mt -> multiply by 32,150.7 (troy oz per metric tonne).
PINK_SHEET_TO_CANONICAL = {
    '$/bbl':     1.0,
    '$/mmbtu':   1.0,
    '$/mt':      1.0,
    '$/troy oz': 32_150.7,
}


def load_world_bank_pinksheet(path: Path) -> pd.DataFrame:
    """Read the WB Pink Sheet annual price file and return long-form."""
    if not path.exists():
        logger.warning(f'PinkSheet not found at {path}; returning empty frame.')
        return pd.DataFrame(columns=['Country', 'Year', 'Resource', 'Metric', 'Value'])

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if 'Annual Prices (Nominal)' not in wb.sheetnames:
        logger.warning(f'PinkSheet: expected sheet "Annual Prices (Nominal)" not found.')
        return pd.DataFrame(columns=['Country', 'Year', 'Resource', 'Metric', 'Value'])

    ws = wb['Annual Prices (Nominal)']

    # Header row is row 7 (1-indexed). Find columns matching PINK_SHEET_MAP.
    header_row = next(ws.iter_rows(min_row=7, max_row=7, values_only=True))
    name_to_col = {}
    for col_idx, name in enumerate(header_row, start=1):
        if name and name in PINK_SHEET_MAP:
            name_to_col[name] = col_idx
    missing = set(PINK_SHEET_MAP) - set(name_to_col)
    if missing:
        logger.warning(f'PinkSheet: {len(missing)} columns not found: {sorted(missing)}')

    rows = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        year_val = row[0]
        if not isinstance(year_val, (int, float)):
            continue
        year = int(year_val)

        for pink_name, col_idx in name_to_col.items():
            v = row[col_idx - 1]
            if v is None or (isinstance(v, str) and v.strip() in ('', '..', '...')):
                continue
            try:
                v = float(v)
            except (TypeError, ValueError):
                continue

            resource, native_unit = PINK_SHEET_MAP[pink_name]
            conv = PINK_SHEET_TO_CANONICAL.get(native_unit, 1.0)
            value_canonical = v * conv

            rows.append({
                'Country':  'World',
                'Year':     year,
                'Resource': resource,
                'Metric':   'Price',
                'Value':    value_canonical,
            })

    df = pd.DataFrame(rows)
    if len(df) == 0:
        logger.warning('PinkSheet: produced 0 rows. Check file structure.')
        return df

    df = df[(df['Year'] >= cfg.YEAR_MIN) & (df['Year'] <= cfg.YEAR_MAX)]

    # Diagnostic: report by resource group, in canonical units
    grp_summary = (
        df.groupby('Resource')['Value']
        .agg(['count', 'mean'])
        .round(2)
        .sort_index()
    )
    logger.info(
        f'PinkSheet loaded: {len(df):,} rows, '
        f'{df["Resource"].nunique()} resources, '
        f'{df["Year"].min()}-{df["Year"].max()}'
    )
    logger.info(f'PinkSheet per-resource mean (canonical units):\\n{grp_summary.to_string()}')
    return df
'''


def patch():
    with open(NB) as f:
        nb = json.load(f)

    # Find the existing PinkSheet loader cell
    target_idx = None
    for i, c in enumerate(nb['cells']):
        if c['cell_type'] != 'code':
            continue
        src = ''.join(c.get('source', []))
        if 'def load_world_bank_pinksheet' in src:
            target_idx = i
            break

    if target_idx is None:
        print('ERROR: could not find existing load_world_bank_pinksheet cell.')
        sys.exit(1)

    nb['cells'][target_idx]['source'] = NEW_LOADER_SOURCE.splitlines(keepends=True)
    nb['cells'][target_idx]['outputs'] = []
    nb['cells'][target_idx]['execution_count'] = None

    with open(NB, 'w') as f:
        json.dump(nb, f, indent=1, ensure_ascii=False)

    print(f'Replaced PinkSheet loader at cell {target_idx}.')


if __name__ == '__main__':
    patch()
