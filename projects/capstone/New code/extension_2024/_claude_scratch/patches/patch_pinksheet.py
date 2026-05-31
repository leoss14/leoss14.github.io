"""Wire World Bank Pink Sheet into e0_NR_extraction.ipynb.

Three edits:
  1. Insert a new cell after cell 23 with the PinkSheet loader.
  2. Patch cell 9: add 'PinkSheet' to the price priority chains.
  3. Patch cell 31: call the loader and add to combine_and_clean.

Idempotent. Re-running on an already-patched notebook is a no-op.
"""
import json, sys, copy

NB = '/Users/leoss/Desktop/GitHub/leoss14.github.io/projects/capstone/New code/extension_2024/e0_NR_extraction.ipynb'

LOADER_HEADER_MD = '## Source 6 -- World Bank Pink Sheet (annual commodity prices)'
LOADER_CELL_SOURCE = '''"""
World Bank Commodity Price Data (the "Pink Sheet").

Annual nominal USD prices, 1960-present, for crude oil, coal, natural gas,
LNG, plus base and precious metals. Single-sheet long-form output matching
the schema produced by the other source loaders.

Download: https://www.worldbank.org/en/research/commodity-markets
File: CMO-Historical-Data-Annual.xlsx (place in rawdata/).
"""
import openpyxl

PINK_SHEET_PATH = RAW / 'CMO-Historical-Data-Annual.xlsx'

# Map Pink Sheet column header -> (canonical Resource name, native unit).
# Multiple oil benchmarks collapse to a single canonical "Oil" series (using
# the average), same for coal (Australian) and natural gas (US Henry Hub).
# Pink Sheet uses US spelling "Aluminum"; canonical is UK "Aluminium".
PINK_SHEET_MAP = {
    'Crude oil, average':           ('Oil',            '$/bbl'),
    'Coal, Australian':             ('Coal',           '$/mt'),
    'Natural gas, US':              ('Natural Gas',    '$/mmbtu'),
    'Aluminum':                     ('Aluminium',      '$/mt'),
    'Copper':                       ('Copper',         '$/mt'),
    'Lead':                         ('Lead',           '$/mt'),
    'Tin':                          ('Tin',            '$/mt'),
    'Nickel':                       ('Nickel',         '$/mt'),
    'Zinc':                         ('Zinc',           '$/mt'),
    'Iron ore, cfr spot':           ('Iron ore',       '$/dmtu'),
    'Gold':                         ('Gold',           '$/troy oz'),
    'Platinum':                     ('Platinum Group', '$/troy oz'),
    'Silver':                       ('Silver',         '$/troy oz'),
}

# Unit conversion to $/mt where the native unit isn't already $/mt.
# Conversion factor x such that price_native * x = $/mt.
# These match the unit basis used by combine_and_clean for other sources.
PINK_SHEET_UNIT_CONVERSION = {
    '$/bbl':     7.33,        # ~7.33 barrels per metric tonne of crude (varies by API gravity)
    '$/mmbtu':   52.0,        # ~52 mmbtu per metric tonne of LNG (industry rule of thumb)
    '$/troy oz': 32_150.7,    # 32,150.7 troy oz per metric tonne
    '$/dmtu':    100.0,       # 1 dmtu = 0.01 tonne of iron content
    '$/mt':      1.0,
}


def load_world_bank_pinksheet(path: Path) -> pd.DataFrame:
    """
    Read the World Bank Pink Sheet annual price file and return long-form.

    Schema: Country='World', Year=int, Resource=canonical name,
            Metric='Price', Value=float (in $/mt where convertible).
    """
    if not path.exists():
        logger.warning(f'PinkSheet not found at {path}; returning empty frame.')
        return pd.DataFrame(columns=['Country', 'Year', 'Resource', 'Metric', 'Value'])

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if 'Annual Prices (Nominal)' not in wb.sheetnames:
        logger.warning(f'PinkSheet: expected sheet "Annual Prices (Nominal)" not found.')
        return pd.DataFrame(columns=['Country', 'Year', 'Resource', 'Metric', 'Value'])

    ws = wb['Annual Prices (Nominal)']

    # Locate header row (row 7) and find columns that match PINK_SHEET_MAP keys.
    header_row = next(ws.iter_rows(min_row=7, max_row=7, values_only=True))
    name_to_col = {}
    for col_idx, name in enumerate(header_row, start=1):
        if name and name in PINK_SHEET_MAP:
            name_to_col[name] = col_idx
    missing = set(PINK_SHEET_MAP) - set(name_to_col)
    if missing:
        logger.warning(f'PinkSheet: {len(missing)} expected columns not found: {sorted(missing)}')

    rows = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        year_val = row[0]
        if not isinstance(year_val, (int, float)):
            continue
        year = int(year_val)

        for pink_name, col_idx in name_to_col.items():
            v = row[col_idx - 1]
            if v is None or (isinstance(v, str) and v.strip() in ('', '..', '...')):
                continue
            try:
                v = float(v)
            except (TypeError, ValueError):
                continue

            resource, native_unit = PINK_SHEET_MAP[pink_name]
            conv = PINK_SHEET_UNIT_CONVERSION.get(native_unit, 1.0)
            price_per_mt = v * conv

            rows.append({
                'Country':  'World',
                'Year':     year,
                'Resource': resource,
                'Metric':   'Price',
                'Value':    price_per_mt,
            })

    df = pd.DataFrame(rows)
    if len(df) == 0:
        logger.warning('PinkSheet: produced 0 rows. Check file structure.')
        return df

    # Clip to configured year range
    df = df[(df['Year'] >= cfg.YEAR_MIN) & (df['Year'] <= cfg.YEAR_MAX)]
    logger.info(
        f'PinkSheet loaded: {len(df):,} rows, '
        f'{df["Resource"].nunique()} resources, '
        f'{df["Year"].min()}-{df["Year"].max()}'
    )
    return df
'''


def patch_notebook():
    with open(NB) as f:
        nb = json.load(f)

    # ── Edit 1: insert PinkSheet loader cell ──
    # Find an existing cell that mentions PinkSheet (idempotency check)
    already_inserted = any(
        'load_world_bank_pinksheet' in ''.join(c.get('source', []))
        for c in nb['cells']
    )

    if already_inserted:
        print('PinkSheet loader already present. Skipping insertion.')
    else:
        # Locate the "Combine and clean" markdown header — insert before it
        insert_idx = None
        for i, c in enumerate(nb['cells']):
            if c['cell_type'] == 'markdown':
                txt = ''.join(c.get('source', []))
                if '## Combine and clean' in txt:
                    insert_idx = i
                    break
        if insert_idx is None:
            print('ERROR: could not locate "## Combine and clean" section to insert before.')
            sys.exit(1)

        md_cell = {
            'cell_type': 'markdown',
            'metadata': {},
            'source': [LOADER_HEADER_MD],
        }
        py_cell = {
            'cell_type': 'code',
            'metadata': {},
            'execution_count': None,
            'outputs': [],
            'source': LOADER_CELL_SOURCE.splitlines(keepends=True),
        }
        nb['cells'].insert(insert_idx, md_cell)
        nb['cells'].insert(insert_idx + 1, py_cell)
        print(f'Inserted PinkSheet loader cells at index {insert_idx} (md) and {insert_idx+1} (code).')

    # ── Edit 2: patch source priority lists in cell 9 ──
    # Find by content (cell index may have shifted by the insertion)
    cell9_idx = None
    for i, c in enumerate(nb['cells']):
        src = ''.join(c.get('source', []))
        if c['cell_type'] == 'code' and 'RESOURCE_PRIORITY' in src and '_PRICE' in src:
            cell9_idx = i
            break
    if cell9_idx is None:
        print('ERROR: could not locate the source priority cell (RESOURCE_PRIORITY).')
        sys.exit(1)

    pri_src = ''.join(nb['cells'][cell9_idx]['source'])

    PATCHES = [
        # _PRICE chain: insert PinkSheet after ConsolidatedPrices
        (
            "_PRICE             = ['ConsolidatedPrices', 'USGS', 'EI_Prices', 'OWID', 'EI_Excel', 'EI_CSV']",
            "_PRICE             = ['ConsolidatedPrices', 'PinkSheet', 'USGS', 'EI_Prices', 'OWID', 'EI_Excel', 'EI_CSV']",
        ),
        # Oil Price: add PinkSheet between EI_Prices and GasPrice
        (
            "('Oil',          'Price'):       ['ConsolidatedPrices', 'EI_Prices', 'GasPrice', 'OWID', 'EI_Excel', 'EI_CSV'],",
            "('Oil',          'Price'):       ['ConsolidatedPrices', 'EI_Prices', 'PinkSheet', 'GasPrice', 'OWID', 'EI_Excel', 'EI_CSV'],",
        ),
        # Natural Gas Price
        (
            "('Natural Gas',  'Price'):       ['ConsolidatedPrices', 'GasPrice', 'EI_Prices', 'OWID', 'EI_Excel', 'EI_CSV'],",
            "('Natural Gas',  'Price'):       ['ConsolidatedPrices', 'GasPrice', 'PinkSheet', 'EI_Prices', 'OWID', 'EI_Excel', 'EI_CSV'],",
        ),
        # Coal Price
        (
            "('Coal',         'Price'):       ['ConsolidatedPrices', 'EI_Prices', 'GasPrice', 'USGS', 'OWID', 'EI_Excel', 'EI_CSV'],",
            "('Coal',         'Price'):       ['ConsolidatedPrices', 'EI_Prices', 'PinkSheet', 'GasPrice', 'USGS', 'OWID', 'EI_Excel', 'EI_CSV'],",
        ),
    ]

    applied = 0
    for old, new in PATCHES:
        if old in pri_src:
            pri_src = pri_src.replace(old, new)
            applied += 1
        elif new in pri_src:
            pass  # already patched
        else:
            print(f'WARNING: priority patch not found:\n  {old[:80]}...')

    if applied > 0:
        nb['cells'][cell9_idx]['source'] = pri_src.splitlines(keepends=True)
        nb['cells'][cell9_idx]['outputs'] = []
        nb['cells'][cell9_idx]['execution_count'] = None
        print(f'Cell {cell9_idx} (priority): applied {applied} patches.')
    elif all(new in pri_src for _, new in PATCHES):
        print(f'Cell {cell9_idx} (priority): already fully patched.')

    # ── Edit 3: patch the run cell (cell 31) ──
    # Find by content
    run_idx = None
    for i, c in enumerate(nb['cells']):
        src = ''.join(c.get('source', []))
        if c['cell_type'] == 'code' and 'combine_and_clean' in src and 'load_ei_csv' in src:
            run_idx = i
            break
    if run_idx is None:
        print('ERROR: could not locate the run cell (combine_and_clean call).')
        sys.exit(1)

    run_src = ''.join(nb['cells'][run_idx]['source'])

    RUN_PATCHES = [
        # Add the loader call right after `consol = load_consolidated_prices(...)`
        (
            "consol      = load_consolidated_prices(CONSOL_PRICE_PATH)  # consolidated price file",
            "consol      = load_consolidated_prices(CONSOL_PRICE_PATH)  # consolidated price file\n"
            "pinksheet   = load_world_bank_pinksheet(PINK_SHEET_PATH)   # World Bank annual commodity prices",
        ),
        # Add to the source list in combine_and_clean
        (
            "    [ei_csv, ei_minerals, ei_prices, owid, usgs, gas, consol],\n"
            "    source_labels=['EI_CSV', 'EI_Excel', 'EI_Prices', 'OWID', 'USGS', 'GasPrice', 'ConsolidatedPrices'],",
            "    [ei_csv, ei_minerals, ei_prices, owid, usgs, gas, consol, pinksheet],\n"
            "    source_labels=['EI_CSV', 'EI_Excel', 'EI_Prices', 'OWID', 'USGS', 'GasPrice', 'ConsolidatedPrices', 'PinkSheet'],",
        ),
    ]

    applied_run = 0
    for old, new in RUN_PATCHES:
        if old in run_src:
            run_src = run_src.replace(old, new)
            applied_run += 1
        elif new in run_src:
            pass
        else:
            print(f'WARNING: run patch not found:\n  {old[:80]}...')

    if applied_run > 0:
        nb['cells'][run_idx]['source'] = run_src.splitlines(keepends=True)
        nb['cells'][run_idx]['outputs'] = []
        nb['cells'][run_idx]['execution_count'] = None
        print(f'Cell {run_idx} (run): applied {applied_run} patches.')

    # Write back
    with open(NB, 'w') as f:
        json.dump(nb, f, indent=1, ensure_ascii=False)

    print('Done.')


if __name__ == '__main__':
    patch_notebook()
