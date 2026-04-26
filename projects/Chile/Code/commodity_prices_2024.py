"""
commodity_prices_2024.py
------------------------
Reads 2024 commodity prices directly from:
  COCHILCO Anuario de Estadisticas del Cobre y otros Minerales 2005-2024

Tables parsed:
  Tabla 96 — Copper (¢US$/lb)
  Tabla 97 — Other Metals (mixed units)
  Tabla 98 — Non-Metallic Minerals (US$/MT, min/max bands)

All prices converted to:
  USD/MT  for bulk commodities
  USD/kg  for precious metals (Gold, Silver, Platinum)

Table 98 price bands: midpoint of min/max used throughout.
Minerals with no 2024 data (n.d.) are excluded.

Usage:
  from commodity_prices_2024 import COMMODITY_PRICES_2024, COMMODITY_PRICES_2024_PER_KG
  # or run directly for a summary printout
"""

import os
import openpyxl

# ── Unit conversion constants ──────────────────────────────────────────
LB_TO_MT = 2204.62   # pounds per metric ton
OZ_TO_KG = 32.1507   # troy oz per kg

ANUARIO_PATH = os.environ.get(
    'COCHILCO_ANUARIO_PATH',
    '/Users/leoss/Desktop/Website-/Portfolio/Website-/projects/Chile/Data/'
    'Anuario-de-Estadisticas-del-Cobre-y-otros-Minerales-2005-2024.xlsx'
)


def _load_workbook(path):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"COCHILCO Anuario not found at:\n  {path}\n"
            "Set the COCHILCO_ANUARIO_PATH environment variable to override."
        )
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _find_col(row, value):
    return row.index(value)


# ── Table 96: Copper ──────────────────────────────────────────────────
# COCHILCO Anuario is a Spanish publication; row labels may be in Spanish
# or English depending on the edition.  Both variants are checked.
_T96_SECTION_LABELS = ('Current Dollars', 'Dólares Corrientes', 'Dolares Corrientes')
_T96_LME_LABELS     = ('LME Refined', 'LME Refinado', 'Grado A')

def _parse_t96(wb):
    """Returns LME copper price in cents US$/lb for 2024 (current dollars).

    Handles both English ('Current Dollars', 'LME Refined') and Spanish
    ('Dólares Corrientes', 'LME Refinado') row labels.
    """
    rows = list(wb['Tabla 96'].iter_rows(values_only=True))
    year_row = next(r for r in rows if 2024 in r)
    col = _find_col(year_row, 2024)

    in_current = False
    for row in rows:
        cell = str(row[0]) if row[0] else ''
        if any(lbl in cell for lbl in _T96_SECTION_LABELS):
            in_current = True
        if in_current and any(lbl in cell for lbl in _T96_LME_LABELS):
            try:
                return float(row[col])
            except (TypeError, ValueError):
                continue   # keep scanning in case this row is a header
    raise ValueError(
        "Copper LME 2024 price not found in Tabla 96.\n"
        "Checked section labels: " + str(_T96_SECTION_LABELS) + "\n"
        "Checked row labels:     " + str(_T96_LME_LABELS) + "\n"
        "If the workbook uses different labels, add them to those tuples."
    )


# ── Table 97: Other Metals ────────────────────────────────────────────
def _parse_t97(wb):
    """Returns {mineral_name: (raw_value, unit_string)} for 2024."""
    MINERAL_MAP = {
        'ORO':                'Gold',
        'PLATA':              'Silver',
        'ALUMINIO':           'Aluminum',
        'NÍQUEL':             'Nickel',
        'PLOMO':              'Lead',
        'ESTAÑO':             'Tin',
        'ZINC':               'Zinc',
        'ÓXIDO DE MOLIBDENO': 'Molybdenum',
        'PLATINO':            'Platinum',
    }
    # For minerals with multiple exchange rows, take this source first
    PREFERRED_SOURCE = {
        'Gold':       'HANDY',
        'Silver':     'HANDY',
        'Aluminum':   'CONTADO',
        'Nickel':     'COMERCIANTES',
        'Lead':       'BML',
        'Tin':        'COMERCIANTES',
        'Zinc':       'BML',
        'Molybdenum': 'COMERCIANTES',
        'Platinum':   'LONDON',
    }

    rows = list(wb['Tabla 97'].iter_rows(values_only=True))
    year_row = next(r for r in rows if 2024 in r)
    col = _find_col(year_row, 2024)

    results = {}
    current_mineral = None

    for row in rows:
        cell = str(row[0]).strip() if row[0] else ''

        for spanish, english in MINERAL_MAP.items():
            if cell.upper().startswith(spanish):
                current_mineral = english
                break

        if current_mineral and row[1] and row[col] is not None:
            try:
                val = float(row[col])
            except (TypeError, ValueError):
                continue
            unit = str(row[1])
            source_hint = PREFERRED_SOURCE.get(current_mineral, '')
            if current_mineral not in results or source_hint in cell.upper():
                results[current_mineral] = (val, unit)

    return results


# ── Table 98: Non-Metallic Minerals ──────────────────────────────────
def _parse_t98(wb):
    """Returns {mineral_name: (min_val, max_val)} for 2024 rows with data.

    Column detection: searches for a header row that contains Mínimo/Máximo
    labels alongside the year 2024, rather than assuming col_max = col_min+1
    (which breaks if a notes column is inserted between min and max).
    """
    LABEL_MAP = {
        'Alúmina calcinada Entrega en Reino Unido': 'Alumina',
        'Carbonato de litio grado técnico':         'Lithium',
        'Yodo crudo':                               'Iodine',
        'Ácido bórico (grado técnico)':             'Boric Acid',
    }

    rows = list(wb['Tabla 98'].iter_rows(values_only=True))

    # Locate Min/Max column headers.
    # The workbook layout has the year numbers in one row and the Min./Max. band
    # labels in the NEXT row (e.g. row N = "…2024 None…", row N+1 = "…Min. Max.…").
    # We search both the year row and the immediately following row.
    col_min, col_max = None, None
    for idx, row in enumerate(rows):
        if 2024 not in row:
            continue
        # Check the year row itself, then the row immediately below it
        candidate_rows = [row]
        if idx + 1 < len(rows):
            candidate_rows.append(rows[idx + 1])
        for cand in candidate_rows:
            cand_strs = [str(v).lower() if v else '' for v in cand]
            min_cands = [i for i, s in enumerate(cand_strs) if 'min' in s]
            max_cands = [i for i, s in enumerate(cand_strs) if 'max' in s]
            if min_cands and max_cands:
                col_min = min_cands[-1]
                col_max = max_cands[-1]
                break
        if col_min is not None:
            break

    if col_min is None:
        # Positional fallback: rightmost 2024 column = min; immediate right = max.
        year_row = next(r for r in rows if 2024 in r)
        positions = [i for i, v in enumerate(year_row) if v == 2024]
        col_min = positions[-1]
        col_max = col_min + 1
        import warnings
        warnings.warn(
            "Tabla 98: header labels 'Mínimo'/'Máximo' not found; "
            "falling back to positional col_min / col_min+1. "
            "Verify the workbook if prices look wrong.",
            stacklevel=3,
        )

    results = {}
    for row in rows:
        if not row[0]:
            continue
        label = str(row[0]).strip().lstrip()
        v_min = row[col_min] if col_min < len(row) else None
        v_max = row[col_max] if col_max < len(row) else None

        if v_min is None or v_min == 'n.d.' or v_max is None or v_max == 'n.d.':
            continue
        try:
            v_min, v_max = float(v_min), float(v_max)
        except (TypeError, ValueError):
            continue

        for key, name in LABEL_MAP.items():
            if key in label:
                results[name] = (v_min, v_max)
                break

    return results


# ── Unit conversion ───────────────────────────────────────────────────
def _to_usd_mt(value, unit):
    u = unit.lower()
    if '¢' in u and 'lb' in u:
        return value / 100 * LB_TO_MT
    if 'us$' in u and 'lb' in u:
        return value * LB_TO_MT
    raise ValueError(f"Cannot convert to USD/MT from unit: {unit}")


def _to_usd_kg(value, unit):
    u = unit.lower()
    if 'us$' in u and ('oz' in u or 'ounce' in u):
        return value * OZ_TO_KG
    raise ValueError(f"Expected oz unit for USD/kg conversion, got: {unit}")


# ── Main builder ──────────────────────────────────────────────────────
PER_KG_MINERALS = {'Gold', 'Silver', 'Platinum'}

SOURCE_LABELS_T97 = {
    'Gold':       'COCHILCO Tabla 97 — Handy & Harman 2024',
    'Silver':     'COCHILCO Tabla 97 — Handy & Harman 2024',
    'Aluminum':   'COCHILCO Tabla 97 — LME Higher Grade 2024',
    'Nickel':     'COCHILCO Tabla 97 — US Dealers 2024',
    'Lead':       'COCHILCO Tabla 97 — LME Settlement 2024',
    'Tin':        'COCHILCO Tabla 97 — US Dealers 2024',
    'Zinc':       'COCHILCO Tabla 97 — LME HG Settlement 2024',
    'Molybdenum': 'COCHILCO Tabla 97 — Molybdenum Oxide, US Dealers 2024',
    'Platinum':   'COCHILCO Tabla 97 — London PM Fix 2024',
}
SOURCE_LABELS_T98 = {
    'Alumina':    'COCHILCO Tabla 98 — Calcined Alumina Del. UK 2024 (midpoint)',
    'Lithium':    'COCHILCO Tabla 98 — Lithium Carbonate Technical Grade USA 2024 (midpoint)',
    'Iodine':     'COCHILCO Tabla 98 — Crude Iodine 2024 (midpoint)',
    'Boric Acid': 'COCHILCO Tabla 98 — Boric Acid Technical Grade 2024 (midpoint)',
}


def build_prices(anuario_path=None):
    """
    Parse the COCHILCO Anuario and return five dicts:
      prices_mt  — {mineral: USD/MT}
      prices_kg  — {mineral: USD/kg}  (precious metals only)
      units      — {mineral: unit string}
      sources    — {mineral: source description}
      bands      — {mineral: (min, max)}  (Table 98 only)
    """
    path = anuario_path or ANUARIO_PATH
    wb   = _load_workbook(path)

    prices_mt, prices_kg, units, sources, bands = {}, {}, {}, {}, {}

    # Copper (Table 96)
    cu_cents_lb         = _parse_t96(wb)
    prices_mt['Copper'] = cu_cents_lb / 100 * LB_TO_MT
    units['Copper']     = 'USD/MT'
    sources['Copper']   = 'COCHILCO Tabla 96 — LME Grade A Settlement 2024'

    # Other metals (Table 97)
    for mineral, (val, unit) in _parse_t97(wb).items():
        sources[mineral] = SOURCE_LABELS_T97.get(mineral, f'COCHILCO Tabla 97 — {mineral} 2024')
        if mineral in PER_KG_MINERALS:
            prices_kg[mineral] = _to_usd_kg(val, unit)
            units[mineral]     = 'USD/kg'
        else:
            prices_mt[mineral] = _to_usd_mt(val, unit)
            units[mineral]     = 'USD/MT'

    # Non-metallics (Table 98)
    for mineral, (v_min, v_max) in _parse_t98(wb).items():
        prices_mt[mineral] = (v_min + v_max) / 2
        units[mineral]     = 'USD/MT'
        sources[mineral]   = SOURCE_LABELS_T98.get(mineral, f'COCHILCO Tabla 98 — {mineral} 2024 (midpoint)')
        bands[mineral]     = (v_min, v_max)

    # Lithium product forms — derived from the Tabla 98 LiCO3 midpoint.
    # Multipliers live in pipeline_constants.LITHIUM_PRICE_MULTIPLIERS.
    # A local fallback is included so this module can be imported standalone
    # (e.g. in a plain Python script without the full pipeline installed).
    if 'Lithium' in prices_mt:
        try:
            from pipeline_constants import LITHIUM_PRICE_MULTIPLIERS as _LI_MULT
        except ImportError:
            _LI_MULT = {
                'Lithium Carbonate': 1.00,
                'Lithium Hydroxide': 1.10,   # ~10 % premium — ESTIMATE
                'Lithium Sulfate':   0.70,   # ~30 % discount — ESTIMATE
            }
        _li_base   = prices_mt['Lithium']
        _li_source = sources.get('Lithium', 'COCHILCO Tabla 98 — Lithium Carbonate 2024 (midpoint)')
        _li_band   = bands.get('Lithium')
        for _form, _mult in _LI_MULT.items():
            prices_mt[_form] = _li_base * _mult
            units[_form]     = 'USD/MT'
            _suffix = f' × {_mult:.2f} [ESTIMATE]' if _mult != 1.0 else ''
            sources[_form]   = _li_source + _suffix
            if _li_band and _mult == 1.0:
                bands[_form] = _li_band   # pass through the min/max band for LiCO3

    wb.close()
    return prices_mt, prices_kg, units, sources, bands


# ── Module-level load ─────────────────────────────────────────────────
COMMODITY_PRICES_2024        = {}
COMMODITY_PRICES_2024_PER_KG = {}
COMMODITY_PRICE_UNITS        = {}
COMMODITY_PRICE_SOURCE       = {}
COMMODITY_PRICE_BANDS        = {}

try:
    (COMMODITY_PRICES_2024,
     COMMODITY_PRICES_2024_PER_KG,
     COMMODITY_PRICE_UNITS,
     COMMODITY_PRICE_SOURCE,
     COMMODITY_PRICE_BANDS) = build_prices()
except FileNotFoundError as e:
    print(f"[commodity_prices_2024] WARNING: {e}")


# ── CLI summary ───────────────────────────────────────────────────────
IN_MODEL = {
    'Copper', 'Molybdenum', 'Gold', 'Silver', 'Zinc', 'Iodine',
    'Lithium', 'Lithium Carbonate', 'Lithium Hydroxide', 'Lithium Sulfate',
}

if __name__ == '__main__':
    print("2024 Commodity Prices — COCHILCO Anuario")
    print(f"{'Mineral':<20} {'Price (USD)':>15}  {'Unit':<8}  {'Band':>25}  Source")
    print("─" * 120)

    all_minerals = list(COMMODITY_PRICES_2024.keys()) + list(COMMODITY_PRICES_2024_PER_KG.keys())
    for mineral in all_minerals:
        tag   = "" if mineral in IN_MODEL else "  (*)"
        price = COMMODITY_PRICES_2024.get(mineral) or COMMODITY_PRICES_2024_PER_KG.get(mineral)
        unit  = COMMODITY_PRICE_UNITS.get(mineral, '')
        band  = ""
        if mineral in COMMODITY_PRICE_BANDS:
            lo, hi = COMMODITY_PRICE_BANDS[mineral]
            band = f"[{lo:,.0f} – {hi:,.0f}]"
        print(f"{mineral+tag:<25} {price:>14,.0f}  {unit:<8}  {band:>25}  {COMMODITY_PRICE_SOURCE.get(mineral,'')}")

    print("\n(*) Not currently modelled in the Chilean supply chain — included for completeness.")
    print("Table 98 prices are midpoints of reported annual min/max price bands.")
